from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, HttpResponseForbidden
from django.utils import timezone
from django.db.models import Sum, Count, Q
from datetime import date, timedelta
from decimal import Decimal
import io
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from django.conf import settings

from coupons.models import FuelCoupon
from fuel_logs.models import FuelLog
from maintenance.models import MaintenanceRecord
from vehicles.models import Vehicle
from generators.models import Generator
from drivers.models import Driver
from vendors.models import Vendor
from audit.models import AuditLog
from accounts.models import Department


# ── Helpers ────────────────────────────────────────────────────────────

def dept_q(user, prefix=""):
    """Return department filter dict for a given queryset prefix."""
    p = f"{prefix}__" if prefix else ""
    if user.is_system_admin or not user.department:
        return {}
    return {f"{p}department": user.department} if prefix else {"department": user.department}


def parse_dates(request):
    """Parse date_from / date_to from request GET, default to current month."""
    today = date.today()
    date_from_str = request.GET.get("date_from", "")
    date_to_str   = request.GET.get("date_to", "")
    try:
        date_from = date.fromisoformat(date_from_str) if date_from_str else today.replace(day=1)
    except ValueError:
        date_from = today.replace(day=1)
    try:
        date_to = date.fromisoformat(date_to_str) if date_to_str else today
    except ValueError:
        date_to = today
    return date_from, date_to


# ── Excel helpers ───────────────────────────────────────────────────────

# Colour palette (matches PDF brown/navy theme)
NAVY   = "0F2044"
NAVY2  = "1A3260"
BROWN  = "6B2D0F"
BROWN2 = "8B3D15"
GOLD   = "C8813A"
GOLD2  = "D99850"
LIGHT  = "F4F6FA"
STRIPE = "EEF1F8"
WHITE  = "FFFFFF"
MUTED  = "6B7A99"


def xl_workbook(title):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    ws.sheet_view.showGridLines = False
    return wb, ws


def _get_logo_image(target_height_px=68):
    """Load the university logo, resize proportionally, return XLImage + dimensions."""
    try:
        logo_path = getattr(settings, "REPORT_LOGO_PATH", None)
        if not logo_path or not os.path.exists(str(logo_path)):
            return None, 0, 0
        pil_img = PILImage.open(str(logo_path)).convert("RGBA")
        orig_w, orig_h = pil_img.size
        ratio = target_height_px / orig_h
        new_w = int(orig_w * ratio)
        pil_img = pil_img.resize((new_w, target_height_px), PILImage.LANCZOS)
        buf = io.BytesIO()
        pil_img.save(buf, format="PNG")
        buf.seek(0)
        xl_img = XLImage(buf)
        xl_img.width  = new_w
        xl_img.height = target_height_px
        return xl_img, new_w, target_height_px
    except Exception:
        return None, 0, 0


def xl_title_block(ws, title, subtitle, date_from, date_to, n_cols=8):
    """
    Header block matching exact user specification:
      Row 1  — navy, empty  (36pt) — logo anchored here, centred
      Row 2  — navy, institution name centred (36pt)
      Row 3  — navy, empty  (18pt) — bottom padding for logo
      Row 4  — brown, report title (26pt)
      Row 5  — light, meta info (16pt)
      Row 6  — spacer (6pt)
      Row 7  — column headers
      Row 8+ — data
    """
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.utils.units import pixels_to_EMU
    import openpyxl.utils as _utils

    last_col = _utils.get_column_letter(n_cols)
    navy_fill  = PatternFill("solid", fgColor=NAVY)
    brown_fill = PatternFill("solid", fgColor=BROWN)
    light_fill = PatternFill("solid", fgColor=LIGHT)

    # ── Rows 1-3: Navy banner ──
    row_heights = {1: 36.0, 2: 36.0, 3: 18.0}
    for r in (1, 2, 3):
        ws.merge_cells(f"A{r}:{last_col}{r}")
        ws.cell(r, 1).fill = navy_fill
        ws.row_dimensions[r].height = row_heights[r]

    # Institution name in row 2, centred
    c = ws.cell(2, 1)
    c.value     = "NORTH-EASTERN UNIVERSITY, GOMBE  •  FLEET MANAGEMENT SYSTEM"
    c.font      = Font(bold=True, size=12, color=WHITE, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ── Logo: centred over rows 1-3 ──
    # Total banner height in px: rows 1+2+3 = 36+36+18 = 90pt → ~90*1.333 = 120px
    # Logo height = 68px leaves ~26px padding top+bottom
    logo, logo_w, logo_h = _get_logo_image(target_height_px=68)
    if logo:
        # Approximate total sheet width in pixels (col width units * 7px/unit)
        col_widths_px = []
        for ci in range(1, n_cols + 1):
            col_letter = _utils.get_column_letter(ci)
            w = ws.column_dimensions[col_letter].width or 8
            col_widths_px.append(w * 7)
        total_w_px = sum(col_widths_px)

        # Centre the logo horizontally
        left_px = max(0, (total_w_px - logo_w) // 2)
        # Small top padding inside row 1
        top_px  = 8

        left_emu = pixels_to_EMU(int(left_px))
        top_emu  = pixels_to_EMU(int(top_px))

        marker = AnchorMarker(col=0, colOff=left_emu, row=0, rowOff=top_emu)
        logo.anchor = OneCellAnchor(_from=marker, ext=None)
        logo.anchor.ext.cx = pixels_to_EMU(logo_w)
        logo.anchor.ext.cy = pixels_to_EMU(logo_h)
        ws.add_image(logo)

    # ── Row 4: Brown — report title ──
    ws.merge_cells(f"A4:{last_col}4")
    c = ws.cell(4, 1)
    c.value     = title.upper()
    c.font      = Font(bold=True, size=14, color=WHITE, name="Calibri")
    c.fill      = brown_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 25.95

    # ── Row 5: Meta bar ──
    ws.merge_cells(f"A5:{last_col}5")
    period = f"{date_from.strftime('%d %b %Y')} — {date_to.strftime('%d %b %Y')}"
    meta   = f"Period: {period}"
    if subtitle:
        meta = f"{subtitle}   |   {meta}"
    meta += "   |   Confidential — For Internal Use Only"
    c = ws.cell(5, 1)
    c.value     = meta
    c.font      = Font(italic=True, size=9, color=MUTED, name="Calibri")
    c.fill      = light_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = Border(bottom=Side(style="thin", color=GOLD))
    ws.row_dimensions[5].height = 16.05

    # ── Row 6: Spacer ──
    ws.row_dimensions[6].height = 6.0


def xl_header_row(ws, cols, row=7):
    """Column header row — navy fill, white bold text, centred."""
    navy_fill  = PatternFill("solid", fgColor=NAVY2)
    white_font = Font(color=WHITE, bold=True, size=10, name="Calibri")
    center     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border     = Border(
        bottom=Side(style="medium", color=GOLD),
        top=Side(style="thin", color=NAVY),
    )
    for col_idx, label in enumerate(cols, 1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.fill      = navy_fill
        cell.font      = white_font
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[row].height = 24
    # Freeze panes so header + title stays visible when scrolling
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def xl_data_row(ws, values, row, shade=False):
    """Data row with alternating stripe, bottom border, and number formatting."""
    fill   = PatternFill("solid", fgColor=STRIPE) if shade else PatternFill("solid", fgColor=WHITE)
    border = Border(bottom=Side(style="thin", color="DDDDDD"))
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.fill      = fill
        cell.border    = border
        cell.font      = Font(size=10, name="Calibri")
        cell.alignment = Alignment(vertical="center", indent=1)
        if isinstance(val, (int, float, Decimal)):
            cell.number_format = '#,##0.00' if isinstance(val, (float, Decimal)) else '#,##0'
            cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row].height = 18


def xl_totals_row(ws, row, n_cols, label_col=1, value_cols=None, values=None):
    """
    Write a gold-accented totals row.
    value_cols — list of 1-based column indices that get a value
    values     — matching list of values
    """
    gold_fill  = PatternFill("solid", fgColor=GOLD)
    navy_font  = Font(bold=True, size=10, color=NAVY, name="Calibri")
    border     = Border(
        top=Side(style="medium", color=BROWN),
        bottom=Side(style="thin", color=BROWN2),
    )
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill   = gold_fill
        cell.border = border
        cell.font   = navy_font
        cell.alignment = Alignment(vertical="center", indent=1)

    ws.cell(row=row, column=label_col).value = "GRAND TOTAL"

    if value_cols and values:
        for col_idx, val in zip(value_cols, values):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row].height = 22


def xl_response(wb, filename):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ── PDF helper ──────────────────────────────────────────────────────────

def pdf_response(html_string, filename):
    # Serve as HTML — browser prints/saves as PDF with full CSS support
    response = HttpResponse(html_string, content_type="text/html; charset=utf-8")
    return response


# ── Dashboard ───────────────────────────────────────────────────────────

@login_required
def dashboard(request):
    user  = request.user
    today = date.today()

    vdept = dept_q(user)
    fdept = dept_q(user, "vehicle")

    # Vehicle-only dept filter (used for vehicle counts, top-vehicles, dept-scoped alerts)
    # NOTE: fdept currently expands to {"vehicle__department": ...} for non-admins
    # which is correct for VEHICLE records. For records that have BOTH vehicle and
    # generator FKs (FuelCoupon, FuelLog, MaintenanceRecord), we need a Q-based
    # widening: vehicle records stay dept-scoped, generator records are
    # organisation-wide.
    if request.user.is_system_admin or not request.user.department:
        asset_scope_q = Q()
    else:
        asset_scope_q = Q(generator__isnull=False) | Q(vehicle__department=request.user.department)

    coupons_today = FuelCoupon.objects.filter(
        asset_scope_q, issue_datetime__date=today
    ).count()
    coupon_value_today = FuelCoupon.objects.filter(
        asset_scope_q,
        issue_datetime__date=today,
        status__in=FuelCoupon.EXPENSE_STATUSES,
    ).aggregate(t=Sum("total_value"))["t"] or 0
    open_coupons = FuelCoupon.objects.filter(
        asset_scope_q,
        status__in=(FuelCoupon.STATUS_APPROVED, FuelCoupon.STATUS_ISSUED),
    ).count()

    month_start = today.replace(day=1)

    # Fuel and maintenance spend — broken down by asset type so the dashboard
    # can show "₦X total (V: ₦Y, G: ₦Z)" rather than a single opaque figure.
    fuel_logs_month  = FuelLog.objects.filter(asset_scope_q, fuel_date__gte=month_start)
    maint_recs_month = MaintenanceRecord.objects.filter(asset_scope_q, service_date__gte=month_start)

    fuel_spend_veh  = fuel_logs_month.filter(generator__isnull=True).aggregate(t=Sum("actual_cost"))["t"] or 0
    fuel_spend_gen  = fuel_logs_month.filter(generator__isnull=False).aggregate(t=Sum("actual_cost"))["t"] or 0
    maint_spend_veh = maint_recs_month.filter(generator__isnull=True).aggregate(t=Sum("total_cost"))["t"] or 0
    maint_spend_gen = maint_recs_month.filter(generator__isnull=False).aggregate(t=Sum("total_cost"))["t"] or 0

    fuel_spend  = fuel_spend_veh  + fuel_spend_gen
    maint_spend = maint_spend_veh + maint_spend_gen
    monthly_spend = fuel_spend + maint_spend
    monthly_spend_veh = fuel_spend_veh + maint_spend_veh
    monthly_spend_gen = fuel_spend_gen + maint_spend_gen

    active_vehicles = Vehicle.objects.filter(status=Vehicle.STATUS_ACTIVE, **vdept).count()
    total_vehicles  = Vehicle.objects.filter(**vdept).count()

    # Generators are organisation-wide so they're never dept-scoped.
    from generators.models import Generator
    active_generators = Generator.objects.filter(status=Generator.STATUS_ACTIVE).count()
    total_generators  = Generator.objects.count()

    stats = {
        "coupons_today": coupons_today,
        "coupon_value_today": coupon_value_today,
        "open_coupons": open_coupons,
        "monthly_spend": monthly_spend,
        "monthly_spend_veh": monthly_spend_veh,
        "monthly_spend_gen": monthly_spend_gen,
        "active_vehicles": active_vehicles,
        "total_vehicles": total_vehicles,
        "active_generators": active_generators,
        "total_generators": total_generators,
    }

    alerts = []
    user = request.user

    # ── Service due alerts (only if user can see maintenance) ──
    service_due = []
    if user.has_module_perm("maintenance", "read"):
        cutoff = today + timedelta(days=14)
        service_due_qs = MaintenanceRecord.objects.filter(
            asset_scope_q,
            next_service_date__lte=cutoff, next_service_date__gte=today,
        ).select_related("vehicle", "generator").order_by("next_service_date")
        seen_v, seen_g = set(), set()
        for rec in service_due_qs:
            # Only one alert per asset (the next upcoming service)
            if rec.is_for_vehicle:
                if rec.vehicle_id in seen_v: continue
                seen_v.add(rec.vehicle_id)
                service_due.append(rec)
                alerts.append({
                    "type": "warning",
                    "title": f"Service due: {rec.vehicle.plate_number}",
                    "detail": f"{rec.vehicle.make} {rec.vehicle.model} — due {rec.next_service_date}",
                    "url": f"/vehicles/{rec.vehicle.pk}/",
                })
            elif rec.is_for_generator:
                if rec.generator_id in seen_g: continue
                seen_g.add(rec.generator_id)
                service_due.append(rec)
                alerts.append({
                    "type": "warning",
                    "title": f"Service due: {rec.generator.tag}",
                    "detail": f"{rec.generator.name} — due {rec.next_service_date}",
                    "url": f"/generators/{rec.generator.pk}/",
                })

    # ── Driver licence alerts (only if user can see drivers) ──
    license_alerts = []
    if user.has_module_perm("drivers", "read"):
        license_cutoff = today + timedelta(days=30)
        license_alerts = Driver.objects.filter(
            status=Driver.STATUS_ACTIVE,
            license_expiry__lte=license_cutoff,
        ).order_by("license_expiry")
        for d in license_alerts:
            alerts.append({
                "type": "danger",
                "title": f"License expiring: {d.full_name}",
                "detail": f"Expires {d.license_expiry} ({d.days_until_license_expiry} days left)",
                "url": f"/drivers/{d.pk}/",
            })

    fuel_map  = {r["vehicle_id"]: r["fuel"]  for r in FuelLog.objects.filter(fuel_date__gte=month_start, vehicle__isnull=False, **fdept).values("vehicle_id").annotate(fuel=Sum("actual_cost"))}
    maint_map = {r["vehicle_id"]: r["maint"] for r in MaintenanceRecord.objects.filter(service_date__gte=month_start, vehicle__isnull=False, **fdept).values("vehicle_id").annotate(maint=Sum("total_cost"))}
    all_ids   = set(fuel_map) | set(maint_map)
    top_raw   = sorted([{"vehicle_id": v, "total_cost": (fuel_map.get(v) or 0) + (maint_map.get(v) or 0)} for v in all_ids], key=lambda x: x["total_cost"], reverse=True)
    vehicles_map = {v.pk: v for v in Vehicle.objects.filter(pk__in=[r["vehicle_id"] for r in top_raw[:5]])}
    top_vehicles = []
    for r in top_raw[:5]:
        v = vehicles_map.get(r["vehicle_id"])
        if v:
            v.total_cost = r["total_cost"]
            top_vehicles.append(v)

    recent_activity = AuditLog.objects.select_related("user")[:10]

    # ── Fleet licence alert (only if user can see vehicles) ──
    from vehicles.models import FleetLicenceExpiry, MonthlyFuelDismissal
    fleet_licence = FleetLicenceExpiry.get()
    if fleet_licence and user.has_module_perm("vehicles", "read") and (fleet_licence.is_expired or fleet_licence.is_expiring_soon):
        alerts.append({
            "type":   "danger" if fleet_licence.is_expired else "warning",
            "title":  "Fleet vehicle licences expired" if fleet_licence.is_expired else "Fleet vehicle licences expiring soon",
            "detail": f"Expiry: {fleet_licence.expiry_date} ({fleet_licence.days_until_expiry} days left)" if not fleet_licence.is_expired else f"Expired on {fleet_licence.expiry_date}",
            "url":    "/vehicles/",
        })

    # ── Monthly fuel reminders ──
    monthly_fuel_vehicles = []
    if today.day >= 1:  # Show from start of month
        dismissed_ids = MonthlyFuelDismissal.objects.filter(
            month=today.month, year=today.year
        ).values_list("vehicle_id", flat=True)
        monthly_fuel_vehicles = Vehicle.objects.filter(
            needs_monthly_fuel=True, status=Vehicle.STATUS_ACTIVE, **vdept
        ).exclude(id__in=dismissed_ids).select_related("department")

    return render(request, "dashboard/dashboard.html", {
        "stats": stats,
        "alerts": alerts,
        "service_due": service_due,
        "license_alerts": license_alerts,
        "fleet_licence": fleet_licence,
        "monthly_fuel_vehicles": monthly_fuel_vehicles,
        "top_vehicles": top_vehicles,
        "recent_activity": recent_activity,
        # Permission flags for dynamic dashboard
        "can_see_coupons":     request.user.has_module_perm("coupons", "read"),
        "can_see_fuel_logs":   request.user.has_module_perm("fuel_logs", "read"),
        "can_see_maintenance": request.user.has_module_perm("maintenance", "read"),
        "can_see_vehicles":    request.user.has_module_perm("vehicles", "read"),
        "can_see_generators":  request.user.has_module_perm("generators", "read"),
        "can_see_drivers":     request.user.has_module_perm("drivers", "read"),
        "can_see_reports":     request.user.has_module_perm("reports", "read"),
        "can_see_audit":       request.user.has_module_perm("audit", "read"),
        "can_edit_vehicles":   request.user.has_module_perm("vehicles", "edit"),
        "can_issue_coupons":   request.user.has_module_perm("coupons", "write"),
    })


# ── Reports Index ────────────────────────────────────────────────────────

@login_required
def reports_index(request):
    if not request.user.has_module_perm("reports", "read"):
        return HttpResponseForbidden()
    return render(request, "reports/index.html")


# ═══════════════════════════════════════════════════════════════════════
# REPORT 1 — Per-Vehicle Spending
# ═══════════════════════════════════════════════════════════════════════

@login_required
def vehicle_spending(request):
    if not request.user.has_module_perm("reports", "read"):
        return HttpResponseForbidden()

    date_from, date_to = parse_dates(request)
    vdept = dept_q(request.user)
    fdept = dept_q(request.user, "vehicle")

    vehicle_filter = request.GET.get("vehicle", "")
    vehicles = Vehicle.objects.filter(**vdept).order_by("plate_number")
    filtered_vehicles = vehicles.filter(pk=vehicle_filter) if vehicle_filter else vehicles

    rows = []
    for v in filtered_vehicles:
        fuel  = FuelLog.objects.filter(vehicle=v, fuel_date__range=[date_from, date_to]).aggregate(
            litres=Sum("actual_litres"), cost=Sum("actual_cost"), count=Count("id"))
        maint = MaintenanceRecord.objects.filter(vehicle=v, service_date__range=[date_from, date_to]).aggregate(
            cost=Sum("total_cost"), count=Count("id"))
        fuel_cost  = fuel["cost"]  or Decimal("0")
        maint_cost = maint["cost"] or Decimal("0")
        total      = fuel_cost + maint_cost
        if total > 0 or request.GET.get("show_all"):
            rows.append({
                "vehicle":     v,
                "fuel_litres": fuel["litres"] or 0,
                "fuel_cost":   fuel_cost,
                "fuel_count":  fuel["count"],
                "maint_cost":  maint_cost,
                "maint_count": maint["count"],
                "total":       total,
            })
    rows.sort(key=lambda x: x["total"], reverse=True)
    grand_fuel  = sum(r["fuel_cost"]  for r in rows)
    grand_maint = sum(r["maint_cost"] for r in rows)
    grand_total = grand_fuel + grand_maint

    fmt = request.GET.get("format", "")
    if fmt == "excel":
        return _vehicle_spending_excel(rows, date_from, date_to, grand_fuel, grand_maint, grand_total)
    if fmt == "pdf":
        return _vehicle_spending_pdf(request, rows, date_from, date_to, grand_fuel, grand_maint, grand_total)

    return render(request, "reports/vehicle_spending.html", {
        "rows": rows, "date_from": date_from, "date_to": date_to,
        "grand_fuel": grand_fuel, "grand_maint": grand_maint, "grand_total": grand_total,
        "vehicles": vehicles, "vehicle_filter": vehicle_filter,
    })


def _vehicle_spending_excel(rows, df, dt, gf, gm, gt):
    n_cols = 8
    wb, ws = xl_workbook("Vehicle Spending")
    xl_title_block(ws, "Per-Vehicle Spending Report", "All Vehicles", df, dt, n_cols=n_cols)
    cols = ["Plate No.", "Make & Model", "Department", "Fuel Litres", "Fuel Cost (₦)", "Maint. Cost (₦)", "Total Spend (₦)", "Maint. Records"]
    xl_header_row(ws, cols, row=7)
    for i, r in enumerate(rows, 8):
        xl_data_row(ws, [
            r["vehicle"].plate_number,
            f"{r['vehicle'].make} {r['vehicle'].model}",
            r["vehicle"].department.name,
            float(r["fuel_litres"]),
            float(r["fuel_cost"]),
            float(r["maint_cost"]),
            float(r["total"]),
            r["maint_count"],
        ], i, shade=(i % 2 == 0))
    totals_row = len(rows) + 8
    xl_totals_row(ws, totals_row, n_cols, label_col=1,
                  value_cols=[5, 6, 7],
                  values=[float(gf), float(gm), float(gt)])
    col_widths = [15, 24, 22, 14, 18, 18, 18, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return xl_response(wb, f"vehicle_spending_{df}_{dt}.xlsx")


def _vehicle_spending_pdf(request, rows, df, dt, gf, gm, gt):
    return render(request, "reports/pdf/vehicle_spending.html", {
        "rows": rows, "date_from": df, "date_to": dt,
        "grand_fuel": gf, "grand_maint": gm, "grand_total": gt,
    })


# ═══════════════════════════════════════════════════════════════════════
# REPORT 2 — Fleet Spending Summary
# ═══════════════════════════════════════════════════════════════════════

@login_required
def fleet_summary(request):
    if not request.user.has_module_perm("reports", "read"):
        return HttpResponseForbidden()

    date_from, date_to = parse_dates(request)
    fdept = dept_q(request.user, "vehicle")

    fuel_qs  = FuelLog.objects.filter(fuel_date__range=[date_from, date_to], **fdept)
    maint_qs = MaintenanceRecord.objects.filter(service_date__range=[date_from, date_to], **fdept)

    total_fuel_cost   = fuel_qs.aggregate(t=Sum("actual_cost"))["t"] or 0
    total_fuel_litres = fuel_qs.aggregate(t=Sum("actual_litres"))["t"] or 0
    total_maint_cost  = maint_qs.aggregate(t=Sum("total_cost"))["t"] or 0
    total_spend       = total_fuel_cost + total_maint_cost
    total_coupons     = FuelCoupon.objects.filter(
        issue_datetime__date__gte=date_from, issue_datetime__date__lte=date_to,
        **fdept
    ).exclude(status__in=[FuelCoupon.STATUS_REJECTED, FuelCoupon.STATUS_CANCELLED]).count()
    total_maint_recs  = maint_qs.count()

    # By department
    dept_rows = []
    for dept in Department.objects.filter(is_active=True):
        fc = FuelLog.objects.filter(fuel_date__range=[date_from, date_to], vehicle__department=dept).aggregate(t=Sum("actual_cost"))["t"] or 0
        mc = MaintenanceRecord.objects.filter(service_date__range=[date_from, date_to], vehicle__department=dept).aggregate(t=Sum("total_cost"))["t"] or 0
        if fc > 0 or mc > 0:
            dept_rows.append({"dept": dept, "fuel": fc, "maint": mc, "total": fc + mc})
    dept_rows.sort(key=lambda x: x["total"], reverse=True)

    # By service type
    maint_by_type = []
    for val, label in MaintenanceRecord.SERVICE_CHOICES:
        cost = maint_qs.filter(service_type=val).aggregate(t=Sum("total_cost"))["t"] or 0
        if cost > 0:
            maint_by_type.append({"label": label, "cost": cost})
    maint_by_type.sort(key=lambda x: x["cost"], reverse=True)

    ctx = {
        "date_from": date_from, "date_to": date_to,
        "total_fuel_cost": total_fuel_cost, "total_fuel_litres": total_fuel_litres,
        "total_maint_cost": total_maint_cost, "total_spend": total_spend,
        "total_coupons": total_coupons, "total_maint_recs": total_maint_recs,
        "dept_rows": dept_rows, "maint_by_type": maint_by_type,
    }

    fmt = request.GET.get("format", "")
    if fmt == "excel":
        return _fleet_summary_excel(ctx)
    if fmt == "pdf":
        return _fleet_summary_pdf(request, ctx)

    return render(request, "reports/fleet_summary.html", ctx)


def _fleet_summary_excel(ctx):
    df, dt = ctx["date_from"], ctx["date_to"]
    n_cols = 4
    wb, ws = xl_workbook("Fleet Summary")
    xl_title_block(ws, "Fleet Spending Summary", "University-wide overview", df, dt, n_cols=n_cols)
    ws.sheet_view.showGridLines = False

    # KPI block starts at row 8 (after 6-row title block + spacer)
    kpi_header_fill = PatternFill("solid", fgColor=NAVY2)
    ws.cell(8, 1, "SUMMARY METRICS").font = Font(bold=True, size=10, color=WHITE, name="Calibri")
    ws.cell(8, 1).fill = kpi_header_fill
    ws.cell(8, 2, "VALUE").font = Font(bold=True, size=10, color=WHITE, name="Calibri")
    ws.cell(8, 2).fill = kpi_header_fill
    ws.row_dimensions[8].height = 22
    kpis = [
        ("Total Fuel Cost (₦)",       float(ctx["total_fuel_cost"])),
        ("Total Fuel Litres",          float(ctx["total_fuel_litres"] or 0)),
        ("Total Maintenance Cost (₦)", float(ctx["total_maint_cost"])),
        ("Grand Total Spend (₦)",      float(ctx["total_spend"])),
        ("Coupons Issued",             ctx["total_coupons"]),
        ("Maintenance Records",        ctx["total_maint_recs"]),
    ]
    for i, (k, v) in enumerate(kpis, 9):
        shade = (i % 2 == 0)
        fill = PatternFill("solid", fgColor=STRIPE if shade else WHITE)
        c1 = ws.cell(i, 1, k)
        c1.font = Font(size=10, name="Calibri")
        c1.fill = fill
        c1.alignment = Alignment(indent=1, vertical="center")
        c2 = ws.cell(i, 2, v)
        c2.font = Font(size=10, name="Calibri")
        c2.fill = fill
        c2.alignment = Alignment(horizontal="right", vertical="center")
        if isinstance(v, float):
            c2.number_format = '#,##0.00'
        ws.row_dimensions[i].height = 18

    # Department breakdown — starts after KPI block (row 8 header + 6 KPI rows = row 16)
    row = 16
    ws.cell(row, 1, "BREAKDOWN BY DEPARTMENT").font = Font(bold=True, size=10, color=WHITE, name="Calibri")
    ws.cell(row, 1).fill = PatternFill("solid", fgColor=BROWN)
    ws.row_dimensions[row].height = 20
    row += 1
    xl_header_row(ws, ["Department", "Fuel Cost (₦)", "Maint. Cost (₦)", "Total (₦)"], row)
    row += 1
    for i, r in enumerate(ctx["dept_rows"], row):
        xl_data_row(ws, [r["dept"].name, float(r["fuel"]), float(r["maint"]), float(r["total"])], i, shade=(i%2==0))
        row += 1
    xl_totals_row(ws, row, n_cols, label_col=1,
                  value_cols=[2, 3, 4],
                  values=[float(ctx["total_fuel_cost"]), float(ctx["total_maint_cost"]), float(ctx["total_spend"])])

    for col, w in zip(["A","B","C","D"], [30, 20, 20, 20]):
        ws.column_dimensions[col].width = w
    return xl_response(wb, f"fleet_summary_{df}_{dt}.xlsx")


def _fleet_summary_pdf(request, ctx):
    return render(request, "reports/pdf/fleet_summary.html", ctx)


# ═══════════════════════════════════════════════════════════════════════
# REPORT — Generator Spending (mirror of Vehicle Spending)
# ═══════════════════════════════════════════════════════════════════════

@login_required
def generator_spending(request):
    """
    Per-generator spending: fuel + maintenance over a date range.

    Generators are organisation-wide so this is NOT department-scoped.
    Generator coupons live in FuelLog rows with generator__isnull=False;
    generator maintenance in MaintenanceRecord rows with the same predicate.
    """
    if not request.user.has_module_perm("reports", "read"):
        return HttpResponseForbidden()

    date_from, date_to = parse_dates(request)

    # Optional filters
    generator_filter = request.GET.get("generator", "")
    building_filter  = request.GET.get("building", "").strip()
    fuel_type_filter = request.GET.get("fuel_type", "")

    generators = Generator.objects.all().order_by("tag")
    filtered = generators
    if generator_filter:
        filtered = filtered.filter(pk=generator_filter)
    if building_filter:
        filtered = filtered.filter(building__iexact=building_filter)
    if fuel_type_filter:
        filtered = filtered.filter(fuel_type=fuel_type_filter)

    rows = []
    for g in filtered:
        fuel = FuelLog.objects.filter(
            generator=g, fuel_date__range=[date_from, date_to]
        ).aggregate(
            litres=Sum("actual_litres"), cost=Sum("actual_cost"), count=Count("id"),
        )
        maint = MaintenanceRecord.objects.filter(
            generator=g, service_date__range=[date_from, date_to]
        ).aggregate(cost=Sum("total_cost"), count=Count("id"))
        fuel_cost  = fuel["cost"]  or Decimal("0")
        maint_cost = maint["cost"] or Decimal("0")
        total      = fuel_cost + maint_cost
        if total > 0 or request.GET.get("show_all"):
            rows.append({
                "generator":   g,
                "fuel_litres": fuel["litres"] or 0,
                "fuel_cost":   fuel_cost,
                "fuel_count":  fuel["count"],
                "maint_cost":  maint_cost,
                "maint_count": maint["count"],
                "total":       total,
            })
    rows.sort(key=lambda x: x["total"], reverse=True)
    grand_fuel  = sum(r["fuel_cost"]  for r in rows)
    grand_maint = sum(r["maint_cost"] for r in rows)
    grand_total = grand_fuel + grand_maint

    # Distinct buildings for filter dropdown
    buildings = (
        Generator.objects.exclude(building="")
        .values_list("building", flat=True).distinct().order_by("building")
    )

    fmt = request.GET.get("format", "")
    if fmt == "excel":
        return _generator_spending_excel(rows, date_from, date_to, grand_fuel, grand_maint, grand_total)
    if fmt == "pdf":
        return _generator_spending_pdf(request, rows, date_from, date_to, grand_fuel, grand_maint, grand_total)

    return render(request, "reports/generator_spending.html", {
        "rows": rows, "date_from": date_from, "date_to": date_to,
        "grand_fuel": grand_fuel, "grand_maint": grand_maint, "grand_total": grand_total,
        "generators": generators, "generator_filter": generator_filter,
        "buildings": buildings, "building_filter": building_filter,
        "fuel_choices": Generator.FUEL_CHOICES, "fuel_type_filter": fuel_type_filter,
    })


def _generator_spending_excel(rows, df, dt, gf, gm, gt):
    n_cols = 8
    wb, ws = xl_workbook("Generator Spending")
    xl_title_block(ws, "Per-Generator Spending Report", "All Generators", df, dt, n_cols=n_cols)
    cols = ["Tag", "Name", "Building", "Fuel Litres", "Fuel Cost (₦)",
            "Maintenance Cost (₦)", "Service Count", "Total (₦)"]
    widths = [14, 28, 20, 13, 16, 18, 14, 16]
    xl_header_row(ws, cols, row=7)
    for i, col_w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = col_w

    r = 8
    for row in rows:
        xl_data_row(ws, [
            row["generator"].tag,
            row["generator"].name,
            row["generator"].building,
            float(row["fuel_litres"] or 0),
            float(row["fuel_cost"] or 0),
            float(row["maint_cost"] or 0),
            row["maint_count"] or 0,
            float(row["total"] or 0),
        ], row=r, shade=(r % 2 == 0))
        r += 1

    xl_totals_row(ws, row=r, n_cols=n_cols, label_col=1, value_cols=[5, 6, 8],
                  values=[float(gf), float(gm), float(gt)])
    return xl_response(wb, f"generator_spending_{df}_{dt}.xlsx")


def _generator_spending_pdf(request, rows, df, dt, gf, gm, gt):
    return render(request, "reports/pdf/generator_spending.html", {
        "rows": rows, "date_from": df, "date_to": dt,
        "grand_fuel": gf, "grand_maint": gm, "grand_total": gt,
    })


# ═══════════════════════════════════════════════════════════════════════
# REPORT 3 — Monthly Expense Report
# ═══════════════════════════════════════════════════════════════════════

@login_required
def monthly_expense(request):
    """
    Combined monthly fuel + maintenance, broken down by asset type:
      vehicle fuel / vehicle maintenance / generator fuel / generator maintenance.

    Vehicle records are still department-scoped via vehicle.department.
    Generator records are organisation-wide (no department scoping),
    consistent with Gen-1/Gen-2 decisions.
    """
    if not request.user.has_module_perm("reports", "read"):
        return HttpResponseForbidden()

    date_from, date_to = parse_dates(request)
    user = request.user

    # Build Q filters that include generator records unconditionally,
    # but constrain vehicle records to the user's department for non-admins.
    if user.is_system_admin or not user.department:
        veh_scope_q = Q()  # no scoping
    else:
        # vehicle records: must match dept; generator records: always allowed
        veh_scope_q = Q(generator__isnull=False) | Q(vehicle__department=user.department)

    fuel_logs = (
        FuelLog.objects
        .filter(veh_scope_q, fuel_date__range=[date_from, date_to])
        .select_related("vehicle", "driver", "generator", "coupon")
        .order_by("fuel_date")
    )

    maint_records = (
        MaintenanceRecord.objects
        .filter(veh_scope_q, service_date__range=[date_from, date_to])
        .select_related("vehicle", "generator", "vendor")
        .order_by("service_date")
    )

    # Four-way breakdown
    veh_fuel  = sum((l.actual_cost for l in fuel_logs    if l.is_for_vehicle),   Decimal("0"))
    gen_fuel  = sum((l.actual_cost for l in fuel_logs    if l.is_for_generator), Decimal("0"))
    veh_maint = sum((r.total_cost  for r in maint_records if r.is_for_vehicle),   Decimal("0"))
    gen_maint = sum((r.total_cost  for r in maint_records if r.is_for_generator), Decimal("0"))

    total_fuel  = veh_fuel  + gen_fuel
    total_maint = veh_maint + gen_maint
    total_spend = total_fuel + total_maint

    # Per-month buckets for the 4-line chart.
    # Key = "YYYY-MM"; value = dict of the four series.
    from collections import defaultdict
    buckets = defaultdict(lambda: {
        "veh_fuel":  Decimal("0"), "gen_fuel":  Decimal("0"),
        "veh_maint": Decimal("0"), "gen_maint": Decimal("0"),
    })

    for l in fuel_logs:
        key = l.fuel_date.strftime("%Y-%m")
        bucket = buckets[key]
        if l.is_for_vehicle:   bucket["veh_fuel"]  += (l.actual_cost or Decimal("0"))
        elif l.is_for_generator: bucket["gen_fuel"] += (l.actual_cost or Decimal("0"))

    for r in maint_records:
        key = r.service_date.strftime("%Y-%m")
        bucket = buckets[key]
        if r.is_for_vehicle:    bucket["veh_maint"]  += (r.total_cost or Decimal("0"))
        elif r.is_for_generator: bucket["gen_maint"] += (r.total_cost or Decimal("0"))

    # Stable, sorted series for the chart.
    months_sorted = sorted(buckets.keys())
    chart_data = {
        "labels":    months_sorted,
        "veh_fuel":  [float(buckets[m]["veh_fuel"])  for m in months_sorted],
        "gen_fuel":  [float(buckets[m]["gen_fuel"])  for m in months_sorted],
        "veh_maint": [float(buckets[m]["veh_maint"]) for m in months_sorted],
        "gen_maint": [float(buckets[m]["gen_maint"]) for m in months_sorted],
    }

    import json as _json
    ctx = {
        "date_from": date_from, "date_to": date_to,
        "fuel_logs": fuel_logs, "maint_records": maint_records,
        # Combined totals (preserved for backward-compat with PDF/Excel templates)
        "total_fuel": total_fuel, "total_maint": total_maint, "total_spend": total_spend,
        # New four-way breakdown
        "veh_fuel":  veh_fuel,  "gen_fuel":  gen_fuel,
        "veh_maint": veh_maint, "gen_maint": gen_maint,
        # Chart data as JSON for inline use
        "chart_data_json": _json.dumps(chart_data),
    }

    fmt = request.GET.get("format", "")
    if fmt == "excel":
        return _monthly_expense_excel(ctx)
    if fmt == "pdf":
        return render(request, "reports/pdf/monthly_expense.html", ctx)

    return render(request, "reports/monthly_expense.html", ctx)


def _monthly_expense_excel(ctx):
    df, dt = ctx["date_from"], ctx["date_to"]
    n_cols = 6
    wb = openpyxl.Workbook()

    # Sheet 1 — Fuel logs
    ws1 = wb.active
    ws1.title = "Fuel Transactions"
    ws1.sheet_view.showGridLines = False
    xl_title_block(ws1, "Monthly Fuel Transactions", "Fuel log detail", df, dt, n_cols=n_cols)
    xl_header_row(ws1, ["Date", "Coupon ID", "Asset", "Driver / Building", "Litres", "Cost (₦)"], 7)
    for i, log in enumerate(ctx["fuel_logs"], 8):
        if log.is_for_vehicle:
            asset_str   = log.vehicle.plate_number
            context_str = log.driver.full_name if log.driver_id else "—"
        else:
            asset_str   = log.generator.tag
            context_str = log.generator.building
        xl_data_row(ws1, [
            str(log.fuel_date), log.coupon.coupon_id,
            asset_str, context_str,
            float(log.actual_litres), float(log.actual_cost),
        ], i, shade=(i%2==0))
    totals_row1 = len(ctx["fuel_logs"]) + 8
    xl_totals_row(ws1, totals_row1, n_cols, label_col=1,
                  value_cols=[5, 6],
                  values=[float(sum(l.actual_litres for l in ctx["fuel_logs"])),
                          float(ctx["total_fuel"])])
    for col, w in zip(["A","B","C","D","E","F"], [14, 24, 14, 22, 12, 16]):
        ws1.column_dimensions[col].width = w

    # Sheet 2 — Maintenance
    ws2 = wb.create_sheet("Maintenance")
    ws2.sheet_view.showGridLines = False
    xl_title_block(ws2, "Monthly Maintenance Records", "Maintenance detail", df, dt, n_cols=n_cols)
    xl_header_row(ws2, ["Date", "Asset", "Service Type", "Vendor", "Cost (₦)", "Approved By"], 7)
    for i, rec in enumerate(ctx["maint_records"], 8):
        asset_str = rec.vehicle.plate_number if rec.is_for_vehicle else rec.generator.tag
        xl_data_row(ws2, [
            str(rec.service_date), asset_str,
            rec.get_service_type_display(), rec.effective_vendor_name,
            float(rec.total_cost), rec.approved_by,
        ], i, shade=(i%2==0))
    totals_row2 = len(ctx["maint_records"]) + 8
    xl_totals_row(ws2, totals_row2, n_cols, label_col=1,
                  value_cols=[5],
                  values=[float(ctx["total_maint"])])
    for col, w in zip(["A","B","C","D","E","F"], [14, 16, 20, 22, 16, 20]):
        ws2.column_dimensions[col].width = w

    return xl_response(wb, f"monthly_expense_{df}_{dt}.xlsx")


# ═══════════════════════════════════════════════════════════════════════
# REPORT 4 — Fuel Coupon Report
# ═══════════════════════════════════════════════════════════════════════

@login_required
def coupon_report(request):
    if not request.user.has_module_perm("reports", "read"):
        return HttpResponseForbidden()

    date_from, date_to = parse_dates(request)
    user = request.user

    if user.is_system_admin or not user.department:
        scope_q = Q()
    else:
        scope_q = Q(generator__isnull=False) | Q(vehicle__department=user.department)

    coupons = (
        FuelCoupon.objects
        .filter(scope_q,
                issue_datetime__date__gte=date_from,
                issue_datetime__date__lte=date_to)
        .select_related("vehicle", "driver", "generator", "fuel_station", "issued_by")
        .order_by("-issue_datetime")
    )

    by_status = {s: coupons.filter(status=s).count() for s, _ in FuelCoupon.STATUS_CHOICES}
    total_value   = coupons.filter(status__in=FuelCoupon.EXPENSE_STATUSES).aggregate(t=Sum("total_value"))["t"] or 0
    redeemed_cost = FuelLog.objects.filter(
        coupon__in=coupons.filter(status=FuelCoupon.STATUS_REDEEMED)
    ).aggregate(t=Sum("actual_cost"))["t"] or 0

    ctx = {
        "date_from": date_from, "date_to": date_to,
        "coupons": coupons, "by_status": by_status,
        "total_value": total_value, "redeemed_cost": redeemed_cost,
    }

    fmt = request.GET.get("format", "")
    if fmt == "excel":
        return _coupon_report_excel(ctx)
    if fmt == "pdf":
        return render(request, "reports/pdf/coupon_report.html", ctx)

    return render(request, "reports/coupon_report.html", ctx)


def _coupon_report_excel(ctx):
    df, dt = ctx["date_from"], ctx["date_to"]
    n_cols = 8
    wb, ws = xl_workbook("Coupon Report")
    ws.sheet_view.showGridLines = False
    xl_title_block(ws, "Fuel Coupon Report", "All issued coupons", df, dt, n_cols=n_cols)
    xl_header_row(ws, ["Coupon ID", "Issued", "Asset", "Driver / Building", "Station", "Litres", "Value (₦)", "Status"], 7)
    for i, c in enumerate(ctx["coupons"], 8):
        if c.is_for_vehicle:
            asset_str   = c.vehicle.plate_number
            context_str = c.driver.full_name if c.driver_id else "—"
        else:
            asset_str   = c.generator.tag
            context_str = c.generator.building
        xl_data_row(ws, [
            c.coupon_id, c.issue_datetime.strftime("%d/%m/%Y %H:%M"),
            asset_str, context_str, c.fuel_station.name,
            float(c.litres), float(c.total_value), c.get_status_display(),
        ], i, shade=(i%2==0))
    totals_row = len(ctx["coupons"]) + 8
    xl_totals_row(ws, totals_row, n_cols, label_col=1,
                  value_cols=[7],
                  values=[float(ctx["total_value"])])
    for col, w in zip(["A","B","C","D","E","F","G","H"], [24, 18, 14, 22, 22, 10, 16, 12]):
        ws.column_dimensions[col].width = w
    return xl_response(wb, f"coupon_report_{df}_{dt}.xlsx")


# ═══════════════════════════════════════════════════════════════════════
# REPORT 5 — Maintenance History
# ═══════════════════════════════════════════════════════════════════════

@login_required
def maintenance_report(request):
    if not request.user.has_module_perm("reports", "read"):
        return HttpResponseForbidden()

    date_from, date_to = parse_dates(request)
    user = request.user

    # Include both vehicle (dept-scoped) and generator (organisation-wide) records
    if user.is_system_admin or not user.department:
        scope_q = Q()
    else:
        scope_q = Q(generator__isnull=False) | Q(vehicle__department=user.department)

    records = (
        MaintenanceRecord.objects
        .filter(scope_q, service_date__range=[date_from, date_to])
        .select_related("vehicle", "generator", "vendor", "created_by")
        .prefetch_related("items")
        .order_by("-service_date")
    )

    total_cost = records.aggregate(t=Sum("total_cost"))["t"] or 0
    by_type    = {}
    for val, label in MaintenanceRecord.SERVICE_CHOICES:
        cost = records.filter(service_type=val).aggregate(t=Sum("total_cost"))["t"] or 0
        if cost > 0:
            by_type[label] = cost

    ctx = {
        "date_from": date_from, "date_to": date_to,
        "records": records, "total_cost": total_cost, "by_type": by_type,
    }

    fmt = request.GET.get("format", "")
    if fmt == "excel":
        return _maintenance_excel(ctx)
    if fmt == "pdf":
        return render(request, "reports/pdf/maintenance_report.html", ctx)

    return render(request, "reports/maintenance_report.html", ctx)


def _maintenance_excel(ctx):
    df, dt = ctx["date_from"], ctx["date_to"]
    n_cols = 8
    wb, ws = xl_workbook("Maintenance History")
    ws.sheet_view.showGridLines = False
    xl_title_block(ws, "Maintenance History Report", "All maintenance records", df, dt, n_cols=n_cols)
    xl_header_row(ws, ["Date", "Asset Kind", "Asset", "Service Type", "Description", "Vendor", "Cost (₦)", "Approved By"], 7)
    for i, r in enumerate(ctx["records"], 8):
        if r.is_for_vehicle:
            asset_kind  = "Vehicle"
            asset_label = f"{r.vehicle.plate_number} ({r.vehicle.make} {r.vehicle.model})"
        else:
            asset_kind  = "Generator"
            asset_label = f"{r.generator.tag} ({r.generator.name})"
        xl_data_row(ws, [
            str(r.service_date), asset_kind, asset_label,
            r.get_service_type_display(), r.items_summary[:80],
            r.effective_vendor_name, float(r.total_cost), r.approved_by,
        ], i, shade=(i%2==0))
    totals_row = len(ctx["records"]) + 8
    xl_totals_row(ws, totals_row, n_cols, label_col=1,
                  value_cols=[7],
                  values=[float(ctx["total_cost"])])
    for col, w in zip(["A","B","C","D","E","F","G","H"], [13, 12, 28, 18, 32, 22, 16, 20]):
        ws.column_dimensions[col].width = w
    return xl_response(wb, f"maintenance_report_{df}_{dt}.xlsx")


# ═══════════════════════════════════════════════════════════════════════
# REPORT 6 — Vendor Spend Report
# ═══════════════════════════════════════════════════════════════════════

@login_required
def vendor_report(request):
    if not request.user.has_module_perm("reports", "read"):
        return HttpResponseForbidden()

    date_from, date_to = parse_dates(request)
    user = request.user

    # Include vehicle records (dept-scoped) AND all generator records.
    if user.is_system_admin or not user.department:
        scope_q = Q()
    else:
        scope_q = Q(generator__isnull=False) | Q(vehicle__department=user.department)

    rows = []
    for vendor in Vendor.objects.all().order_by("name"):
        fuel  = FuelLog.objects.filter(
            scope_q, fuel_station=vendor, fuel_date__range=[date_from, date_to]
        ).aggregate(cost=Sum("actual_cost"), litres=Sum("actual_litres"), count=Count("id"))
        maint = MaintenanceRecord.objects.filter(
            scope_q, vendor=vendor, service_date__range=[date_from, date_to]
        ).aggregate(cost=Sum("total_cost"), count=Count("id"))
        fc = fuel["cost"]  or Decimal("0")
        mc = maint["cost"] or Decimal("0")
        if fc > 0 or mc > 0:
            rows.append({
                "vendor": vendor, "fuel_cost": fc, "fuel_litres": fuel["litres"] or 0,
                "fuel_txn": fuel["count"], "maint_cost": mc, "maint_txn": maint["count"],
                "total": fc + mc,
            })
    rows.sort(key=lambda x: x["total"], reverse=True)
    grand_total = sum(r["total"] for r in rows)

    ctx = {
        "date_from": date_from, "date_to": date_to,
        "rows": rows, "grand_total": grand_total,
    }

    fmt = request.GET.get("format", "")
    if fmt == "excel":
        return _vendor_excel(ctx)
    if fmt == "pdf":
        return render(request, "reports/pdf/vendor_report.html", ctx)

    return render(request, "reports/vendor_report.html", ctx)


def _vendor_excel(ctx):
    df, dt = ctx["date_from"], ctx["date_to"]
    n_cols = 8
    wb, ws = xl_workbook("Vendor Spend")
    ws.sheet_view.showGridLines = False
    xl_title_block(ws, "Vendor Spend Report", "All active vendors", df, dt, n_cols=n_cols)
    xl_header_row(ws, ["Vendor", "Type", "Fuel Cost (₦)", "Litres", "Fuel Txns", "Maint. Cost (₦)", "Maint. Txns", "Total (₦)"], 7)
    for i, r in enumerate(ctx["rows"], 8):
        xl_data_row(ws, [
            r["vendor"].name, r["vendor"].get_type_display(),
            float(r["fuel_cost"]), float(r["fuel_litres"]),
            r["fuel_txn"], float(r["maint_cost"]), r["maint_txn"], float(r["total"]),
        ], i, shade=(i%2==0))
    totals_row = len(ctx["rows"]) + 8
    xl_totals_row(ws, totals_row, n_cols, label_col=1,
                  value_cols=[8],
                  values=[float(ctx["grand_total"])])
    for col, w in zip(["A","B","C","D","E","F","G","H"], [28, 16, 18, 12, 12, 18, 12, 18]):
        ws.column_dimensions[col].width = w
    return xl_response(wb, f"vendor_report_{df}_{dt}.xlsx")


# ── Driver Payments Report ───────────────────────────────────────────────────

@login_required
def driver_payments(request):
    """
    Summary of trip payments per driver in a period.
    Defaults to current calendar month.
    """
    if not request.user.has_module_perm("reports", "read"):
        return HttpResponseForbidden()

    from trips.models import Trip

    today = date.today()
    df = request.GET.get("date_from") or today.replace(day=1).isoformat()
    dt = request.GET.get("date_to")   or today.isoformat()
    driver_filter = request.GET.get("driver", "")
    payment_type  = request.GET.get("payment_type", "")

    # Department scoping via vehicle
    dept_q = {}
    if not request.user.is_system_admin and request.user.department:
        dept_q = {"vehicle__department": request.user.department}

    qs = Trip.objects.filter(
        trip_date__gte=df, trip_date__lte=dt, **dept_q
    ).select_related("driver", "vehicle")

    if driver_filter:
        qs = qs.filter(driver_id=driver_filter)
    if payment_type:
        qs = qs.filter(driver__payment_type=payment_type)

    # Aggregate by driver
    summary = (
        qs.values("driver_id", "driver__full_name", "driver__staff_id", "driver__payment_type")
          .annotate(trip_count=Count("id"), total_paid=Sum("amount_paid"))
          .order_by("-total_paid")
    )

    total_all = qs.aggregate(s=Sum("amount_paid"))["s"] or Decimal("0.00")
    trip_count_all = qs.count()

    drivers = Driver.objects.order_by("full_name")

    return render(request, "reports/driver_payments.html", {
        "summary": summary,
        "trips":   qs.order_by("-trip_date")[:200],  # detail rows for preview
        "total_all": total_all,
        "trip_count_all": trip_count_all,
        "date_from": df, "date_to": dt,
        "driver_filter": driver_filter,
        "payment_type_filter": payment_type,
        "drivers": drivers,
        "payment_choices": Driver.PAY_CHOICES,
    })


# ── Report Scheduling ────────────────────────────────────────────────────────

@login_required
def schedule_list(request):
    if not request.user.has_module_perm("reports", "read"):
        return HttpResponseForbidden()
    from .models import ReportSchedule
    schedules = ReportSchedule.objects.all()
    return render(request, "reports/schedule_list.html", {
        "schedules": schedules,
        "can_write":  request.user.has_module_perm("reports", "write"),
        "can_edit":   request.user.has_module_perm("reports", "edit"),
        "can_delete": request.user.has_module_perm("reports", "delete"),
    })


@login_required
def schedule_create(request):
    if not request.user.has_module_perm("reports", "write"):
        return HttpResponseForbidden()
    from .models import ReportSchedule
    if request.method == "POST":
        name        = request.POST.get("name", "").strip()
        report_type = request.POST.get("report_type", "")
        fmt         = request.POST.get("format", "pdf")
        recipients  = request.POST.get("recipients", "").strip()
        send_day    = request.POST.get("send_day", "1").strip()
        is_active   = request.POST.get("is_active") == "on"

        errors = {}
        if not name:        errors["name"]       = "Name is required."
        if not report_type: errors["report_type"]= "Report type is required."
        if not recipients:  errors["recipients"] = "At least one recipient email is required."
        try:
            send_day = int(send_day)
            if not (1 <= send_day <= 28):
                errors["send_day"] = "Day must be between 1 and 28."
        except ValueError:
            errors["send_day"] = "Enter a valid day number."

        if not errors:
            ReportSchedule.objects.create(
                name=name, report_type=report_type, format=fmt,
                recipients=recipients, send_day=send_day,
                is_active=is_active, created_by=request.user.full_name,
            )
            messages.success(request, f"Schedule '{name}' created.")
            return redirect("reports:schedule_list")

        return render(request, "reports/schedule_form.html", {
            "errors": errors, "post": request.POST,
            "report_choices": ReportSchedule.REPORT_CHOICES,
            "format_choices": ReportSchedule.FORMAT_CHOICES,
        })

    from .models import ReportSchedule
    return render(request, "reports/schedule_form.html", {
        "errors": {}, "post": {},
        "report_choices": ReportSchedule.REPORT_CHOICES,
        "format_choices": ReportSchedule.FORMAT_CHOICES,
    })


@login_required
def schedule_edit(request, pk):
    if not request.user.has_module_perm("reports", "edit"):
        return HttpResponseForbidden()
    from .models import ReportSchedule
    schedule = get_object_or_404(ReportSchedule, pk=pk)

    if request.method == "POST":
        schedule.name        = request.POST.get("name", schedule.name).strip()
        schedule.report_type = request.POST.get("report_type", schedule.report_type)
        schedule.format      = request.POST.get("format", schedule.format)
        schedule.recipients  = request.POST.get("recipients", "").strip()
        schedule.send_day    = int(request.POST.get("send_day", schedule.send_day))
        schedule.is_active   = request.POST.get("is_active") == "on"
        schedule.save()
        messages.success(request, f"Schedule '{schedule.name}' updated.")
        return redirect("reports:schedule_list")

    return render(request, "reports/schedule_form.html", {
        "obj": schedule, "errors": {}, "post": {},
        "report_choices": ReportSchedule.REPORT_CHOICES,
        "format_choices": ReportSchedule.FORMAT_CHOICES,
    })


@login_required
def schedule_delete(request, pk):
    if not request.user.has_module_perm("reports", "delete"):
        return HttpResponseForbidden()
    from .models import ReportSchedule
    schedule = get_object_or_404(ReportSchedule, pk=pk)
    if request.method == "POST":
        name = schedule.name
        schedule.delete()
        messages.success(request, f"Schedule '{name}' deleted.")
    return redirect("reports:schedule_list")


@login_required
def schedule_send_now(request, pk):
    """Send a scheduled report immediately for a custom date range."""
    if not request.user.has_module_perm("reports", "write"):
        return HttpResponseForbidden()
    from .models import ReportSchedule
    schedule = get_object_or_404(ReportSchedule, pk=pk)

    if request.method == "POST":
        date_from_str = request.POST.get("date_from", "").strip()
        date_to_str   = request.POST.get("date_to", "").strip()

        if not date_from_str or not date_to_str:
            messages.error(request, "Both date from and date to are required.")
            return redirect("reports:schedule_list")

        try:
            from datetime import datetime
            df = datetime.strptime(date_from_str, "%Y-%m-%d").date()
            dt = datetime.strptime(date_to_str,   "%Y-%m-%d").date()
            if df > dt:
                messages.error(request, "Date from must be before date to.")
                return redirect("reports:schedule_list")
        except ValueError:
            messages.error(request, "Invalid date format.")
            return redirect("reports:schedule_list")

        # Build and send using the management command helpers
        try:
            data, filename, mime_type = _build_report_instant(schedule, df, dt)
            _send_email_instant(schedule, data, filename, mime_type, df, dt, request.user.full_name)

            schedule.last_sent = timezone.now()
            schedule.save(update_fields=["last_sent"])

            messages.success(
                request,
                f"Report '{schedule.name}' sent to {', '.join(schedule.recipient_list)} "
                f"for period {df.strftime('%d %b %Y')} — {dt.strftime('%d %b %Y')}."
            )
        except Exception as e:
            messages.error(request, f"Failed to send report: {e}")

        return redirect("reports:schedule_list")

    return redirect("reports:schedule_list")


def _build_report_instant(schedule, df, dt):
    """Build report attachment for instant send with a custom date range."""
    from decimal import Decimal
    from vehicles.models import Vehicle
    from coupons.models import FuelCoupon
    from maintenance.models import MaintenanceRecord
    from fuel_logs.models import FuelLog
    from vendors.models import Vendor
    from accounts.models import Department
    from django.db.models import Sum

    rt  = schedule.report_type
    fmt = schedule.format

    def vehicle_rows():
        rows = []
        for v in Vehicle.objects.filter(status="active").select_related("department"):
            fc = FuelCoupon.objects.filter(vehicle=v, status="redeemed", issue_datetime__date__range=[df,dt])
            mc = MaintenanceRecord.objects.filter(vehicle=v, service_date__range=[df,dt])
            fuel_cost   = fc.aggregate(t=Sum("total_value"))["t"] or Decimal(0)
            maint_cost  = mc.aggregate(t=Sum("total_cost"))["t"]  or Decimal(0)
            fuel_litres = fc.aggregate(t=Sum("litres"))["t"]       or Decimal(0)
            if fuel_cost or maint_cost:
                rows.append({"vehicle": v, "fuel_cost": fuel_cost, "maint_cost": maint_cost,
                             "total": fuel_cost+maint_cost, "fuel_litres": fuel_litres,
                             "maint_count": mc.count()})
        return rows

    def generator_rows():
        """
        Per-generator spending for instant-send reports.

        Uses FuelLog.actual_cost for fuel (matches the interactive view at
        /reports/generators/). Generator coupons aren't suitable here because
        a generator coupon's `total_value` reflects issued amount, not actual
        consumption.
        """
        from generators.models import Generator
        rows = []
        for g in Generator.objects.all().order_by("tag"):
            fl = FuelLog.objects.filter(generator=g, fuel_date__range=[df,dt])
            mc = MaintenanceRecord.objects.filter(generator=g, service_date__range=[df,dt])
            fuel_cost   = fl.aggregate(t=Sum("actual_cost"))["t"]   or Decimal(0)
            fuel_litres = fl.aggregate(t=Sum("actual_litres"))["t"] or Decimal(0)
            maint_cost  = mc.aggregate(t=Sum("total_cost"))["t"]    or Decimal(0)
            total       = fuel_cost + maint_cost
            if total:
                rows.append({"generator": g, "fuel_cost": fuel_cost, "maint_cost": maint_cost,
                             "total": total, "fuel_litres": fuel_litres,
                             "fuel_count": fl.count(), "maint_count": mc.count()})
        rows.sort(key=lambda x: x["total"], reverse=True)
        return rows

    if fmt == "xlsx":
        if rt == "vehicle_spending":
            rows = vehicle_rows()
            gf = sum(r["fuel_cost"] for r in rows)
            gm = sum(r["maint_cost"] for r in rows)
            gt = sum(r["total"] for r in rows)
            resp = _vehicle_spending_excel(rows, df, dt, gf, gm, gt)
        elif rt == "generator_spending":
            rows = generator_rows()
            gf = sum(r["fuel_cost"] for r in rows)
            gm = sum(r["maint_cost"] for r in rows)
            gt = sum(r["total"] for r in rows)
            resp = _generator_spending_excel(rows, df, dt, gf, gm, gt)
        elif rt == "fleet_summary":
            coupons_qs = FuelCoupon.objects.filter(status="redeemed", issue_datetime__date__range=[df,dt])
            maint_qs   = MaintenanceRecord.objects.filter(service_date__range=[df,dt])
            tf = coupons_qs.aggregate(t=Sum("total_value"))["t"] or Decimal(0)
            tl = coupons_qs.aggregate(t=Sum("litres"))["t"]       or Decimal(0)
            tm = maint_qs.aggregate(t=Sum("total_cost"))["t"]      or Decimal(0)
            dept_rows = []
            for dept in Department.objects.filter(is_active=True):
                f = FuelCoupon.objects.filter(vehicle__department=dept, status="redeemed", issue_datetime__date__range=[df,dt]).aggregate(t=Sum("total_value"))["t"] or Decimal(0)
                m = MaintenanceRecord.objects.filter(vehicle__department=dept, service_date__range=[df,dt]).aggregate(t=Sum("total_cost"))["t"] or Decimal(0)
                dept_rows.append({"dept": dept, "fuel": f, "maint": m, "total": f+m})
            ctx = {"date_from": df, "date_to": dt, "total_fuel_cost": tf, "total_fuel_litres": tl,
                   "total_maint_cost": tm, "total_spend": tf+tm,
                   "total_coupons": FuelCoupon.objects.filter(issue_datetime__date__range=[df,dt]).count(),
                   "total_maint_recs": maint_qs.count(), "dept_rows": dept_rows}
            resp = _fleet_summary_excel(ctx)
        elif rt == "monthly_expense":
            logs  = FuelLog.objects.filter(fuel_date__range=[df,dt]).select_related("vehicle","driver","coupon")
            recs  = MaintenanceRecord.objects.filter(service_date__range=[df,dt]).select_related("vehicle","vendor")
            tf    = logs.aggregate(t=Sum("actual_cost"))["t"] or Decimal(0)
            tm    = recs.aggregate(t=Sum("total_cost"))["t"]  or Decimal(0)
            ctx   = {"date_from": df, "date_to": dt, "fuel_logs": logs, "maint_records": recs,
                     "total_fuel": tf, "total_maint": tm, "total_spend": tf + tm}
            resp  = _monthly_expense_excel(ctx)
        elif rt == "coupon_report":
            coupons = FuelCoupon.objects.filter(issue_datetime__date__range=[df,dt]).select_related("vehicle","driver","fuel_station")
            ctx = {"date_from": df, "date_to": dt, "coupons": coupons,
                   "total_value": coupons.filter(status__in=FuelCoupon.EXPENSE_STATUSES).aggregate(t=Sum("total_value"))["t"] or Decimal(0)}
            resp = _coupon_report_excel(ctx)
        elif rt == "maintenance":
            recs = MaintenanceRecord.objects.filter(service_date__range=[df,dt]).select_related("vehicle","vendor")
            ctx  = {"date_from": df, "date_to": dt, "records": recs,
                    "total_cost": recs.aggregate(t=Sum("total_cost"))["t"] or Decimal(0)}
            resp = _maintenance_excel(ctx)
        elif rt == "vendor":
            rows2 = []; grand = Decimal(0)
            for v in Vendor.objects.filter(is_active=True):
                fc = FuelCoupon.objects.filter(fuel_station=v, issue_datetime__date__range=[df,dt])
                mc = MaintenanceRecord.objects.filter(vendor=v, service_date__range=[df,dt])
                fc_cost = fc.aggregate(t=Sum("total_value"))["t"] or Decimal(0)
                mc_cost = mc.aggregate(t=Sum("total_cost"))["t"]  or Decimal(0)
                total   = fc_cost + mc_cost
                if total:
                    rows2.append({"vendor": v, "fuel_cost": fc_cost,
                                  "fuel_litres": fc.aggregate(t=Sum("litres"))["t"] or Decimal(0),
                                  "fuel_txn": fc.count(), "maint_cost": mc_cost,
                                  "maint_txn": mc.count(), "total": total})
                    grand += total
            ctx  = {"date_from": df, "date_to": dt, "rows": rows2, "grand_total": grand}
            resp = _vendor_excel(ctx)
        else:
            raise ValueError(f"Unknown report type: {rt}")

        data     = resp.content
        filename = f"{rt}_{df}_{dt}.xlsx"
        mime     = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    else:
        # PDF schedule — render as self-contained HTML with all CSS inlined
        # Opens in any browser and looks identical to the printable report
        from django.template.loader import render_to_string

        template_map = {
            "vehicle_spending":   "reports/pdf/vehicle_spending.html",
            "generator_spending": "reports/pdf/generator_spending.html",
            "fleet_summary":      "reports/pdf/fleet_summary.html",
            "monthly_expense":    "reports/pdf/monthly_expense.html",
            "coupon_report":      "reports/pdf/coupon_report.html",
            "maintenance":        "reports/pdf/maintenance_report.html",
            "vendor":             "reports/pdf/vendor_report.html",
        }
        template = template_map.get(rt)
        if not template:
            raise ValueError(f"PDF export not supported for report type: {rt}")

        # Build context
        ctx = {"date_from": df, "date_to": dt}
        if rt == "vehicle_spending":
            rows = vehicle_rows()
            ctx.update({"rows": rows,
                        "grand_fuel":  sum(r["fuel_cost"]  for r in rows),
                        "grand_maint": sum(r["maint_cost"] for r in rows),
                        "grand_total": sum(r["total"]       for r in rows)})
        elif rt == "generator_spending":
            rows = generator_rows()
            ctx.update({"rows": rows,
                        "grand_fuel":  sum(r["fuel_cost"]  for r in rows),
                        "grand_maint": sum(r["maint_cost"] for r in rows),
                        "grand_total": sum(r["total"]       for r in rows)})
        elif rt == "monthly_expense":
            logs = FuelLog.objects.filter(fuel_date__range=[df,dt]).select_related("vehicle","driver","coupon")
            recs = MaintenanceRecord.objects.filter(service_date__range=[df,dt]).select_related("vehicle","vendor")
            tf = logs.aggregate(t=Sum("actual_cost"))["t"] or Decimal(0)
            tm = recs.aggregate(t=Sum("total_cost"))["t"]  or Decimal(0)
            ctx.update({"fuel_logs": logs, "maint_records": recs,
                        "total_fuel": tf, "total_maint": tm, "total_spend": tf + tm})
        elif rt == "coupon_report":
            coupons = FuelCoupon.objects.filter(issue_datetime__date__range=[df,dt]).select_related("vehicle","driver","fuel_station")
            ctx.update({"coupons": coupons,
                        "total_value": coupons.filter(status__in=FuelCoupon.EXPENSE_STATUSES).aggregate(t=Sum("total_value"))["t"] or Decimal(0)})
        elif rt == "maintenance":
            recs = MaintenanceRecord.objects.filter(service_date__range=[df,dt]).select_related("vehicle","vendor")
            ctx.update({"records": recs,
                        "total_cost": recs.aggregate(t=Sum("total_cost"))["t"] or Decimal(0)})

        html_string = render_to_string(template, ctx)

        # Load main.css and inline it so the file is self-contained
        css_path = os.path.join(str(settings.BASE_DIR), "static", "css", "main.css")
        if not os.path.exists(css_path):
            css_path = os.path.join(str(settings.BASE_DIR), "staticfiles", "css", "main.css")

        if os.path.exists(css_path):
            with open(css_path, "r", encoding="utf-8") as f:
                css_content = f.read()
            # Replace the <link> tag referencing main.css / reports.css with inline <style>
            import re
            html_string = re.sub(
                r'<link[^>]+\.css[^>]*>',
                f'<style>{css_content}</style>',
                html_string
            )

        data     = html_string.encode("utf-8")
        filename = f"{rt}_{df}_{dt}.html"
        mime     = "text/html"

    return data, filename, mime


def _send_email_instant(schedule, data, filename, mime_type, df, dt, sent_by):
    """Send the built report via email with a custom period label."""
    from django.core.mail import EmailMultiAlternatives
    period_label = f"{df.strftime('%d %b %Y')} — {dt.strftime('%d %b %Y')}"

    # Resolve branding for sender + signature
    try:
        from system_settings.models import SystemSettings
        _brand = SystemSettings.get()
        _from  = _brand.email_from or getattr(settings, "DEFAULT_FROM_EMAIL", "fleet@neu.edu.ng")
        _sig   = _brand.institution_name
        _sys   = _brand.system_name
    except Exception:
        _from = getattr(settings, "DEFAULT_FROM_EMAIL", "fleet@neu.edu.ng")
        _sig  = "Fleet Management"
        _sys  = "Fleet Management"

    subject = f"{_sys} — {schedule.get_report_type_display()} ({period_label})"
    body = (
        f"Hello,\n\n"
        f"Please find attached the {schedule.get_report_type_display()} report "
        f"for the period {period_label}.\n\n"
        f"This report was sent manually by {sent_by}.\n\n"
        f"— Fleet Management System\n"
        f"{_sig}"
    )
    from system_settings.mail import get_mail_connection
    msg = EmailMultiAlternatives(
        subject=subject, body=body,
        from_email=_from,
        to=schedule.recipient_list,
        connection=get_mail_connection(),
    )
    msg.attach(filename, data, mime_type)
    msg.send()
